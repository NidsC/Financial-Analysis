"""
Build sector_medians_dashboard.xlsx — an interactive Excel dashboard with:
  - Sheet: DATA          — raw sector medians (all years, all sectors)
  - Sheet: ACROSS_YEARS  — line chart: pick sector + metric, see time series
  - Sheet: ACROSS_SECTORS — bar chart: pick year + metric, see all sectors side-by-side
  - Sheet: CONTROL       — dropdown selectors driving both charts via named ranges

Approach: openpyxl for structure + data validation dropdowns; chart data pulled
from helper tables that use SUMPRODUCT formulas, so Excel re-renders live when
the user changes a dropdown. Charts are bound to those helper ranges.
"""

import sqlite3
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabel

DB_PATH  = os.path.join(os.path.dirname(__file__), "uk_equity_library.db")
OUT_PATH = os.path.join(os.path.dirname(__file__), "sector_medians_dashboard.xlsx")

SECTORS = [
    "Basic Materials", "Communication Services", "Consumer Cyclical",
    "Consumer Defensive", "Energy", "Financial Services", "Healthcare",
    "Industrials", "Real Estate", "Technology", "Utilities",
]
YEARS = [str(y) for y in range(2011, 2027)]

METRICS = {
    "PE Ratio":          "median_pe_ratio",
    "EV/EBITDA":         "median_ev_ebitda",
    "Price-to-Book":     "median_price_to_book",
    "PEG Ratio":         "median_peg_ratio",
    "ROIC":              "median_roic",
    "FCF Yield":         "median_fcf_yield",
    "Dividend Yield":    "median_dividend_yield",
    "Payout Ratio":      "median_payout_ratio",
    "Current Ratio":     "median_current_ratio",
    "ROE":               "median_roe",
    "Debt/Equity":       "median_debt_to_equity",
    "Net Profit Margin": "median_net_profit_margin",
    "Gross Margin":      "median_gross_profit_margin",
    "Revenue Growth":    "median_revenue_growth",
    "Net Income Growth": "median_net_income_growth",
}
METRIC_LABELS = list(METRICS.keys())
METRIC_COLS   = list(METRICS.values())

# ── Colour palette ───────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "BDD7EE"
ACCENT     = "F4B942"
WHITE      = "FFFFFF"
GREY_BG    = "F2F2F2"
BORDER_CLR = "BFBFBF"

SECTOR_COLOURS = [
    "2E75B6","ED7D31","A9D18E","FFC000","4472C4",
    "70AD47","FF0000","7030A0","00B0F0","92D050","FF7F7F",
]


def thin_border(sides="all"):
    s = Side(style="thin", color=BORDER_CLR)
    n = Side(style=None)
    if sides == "all":
        return Border(left=s, right=s, top=s, bottom=s)
    b = Border()
    for side in sides:
        setattr(b, side, s)
    return b


def header_cell(ws, row, col, text, bg=DARK_BLUE, fg=WHITE, bold=True, size=11, align="center"):
    c = ws.cell(row=row, column=col, value=text)
    c.font  = Font(bold=bold, color=fg, size=size)
    c.fill  = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = thin_border()
    return c


def data_cell(ws, row, col, value, bg=WHITE, align="center", num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill  = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = thin_border()
    if num_fmt:
        c.number_format = num_fmt
    return c


# ── Pull data from DB ─────────────────────────────────────────────────────────
def load_data():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT * FROM sector_medians ORDER BY sector, year"
    ).fetchall()
    conn.close()
    return rows


# ── Sheet 1: DATA ─────────────────────────────────────────────────────────────
def build_data_sheet(wb, rows):
    ws = wb.create_sheet("DATA")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C2"

    # Header row
    header_cell(ws, 1, 1, "Sector", bg=DARK_BLUE)
    header_cell(ws, 1, 2, "Year",   bg=DARK_BLUE)
    for ci, label in enumerate(METRIC_LABELS, start=3):
        header_cell(ws, 1, ci, label, bg=DARK_BLUE, size=10)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 8
    for ci in range(3, 3 + len(METRIC_LABELS)):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    # Data rows
    pct_cols = {
        "FCF Yield", "Dividend Yield", "Payout Ratio",
        "ROE", "Net Profit Margin", "Gross Margin",
        "Revenue Growth", "Net Income Growth", "ROIC",
    }
    pct_indices = [i+3 for i, m in enumerate(METRIC_LABELS) if m in pct_cols]

    for ri, row in enumerate(rows, start=2):
        bg = WHITE if ri % 2 == 0 else GREY_BG
        data_cell(ws, ri, 1, row["sector"],  bg=bg, align="left")
        data_cell(ws, ri, 2, int(row["year"]), bg=bg)
        for ci, col in enumerate(METRIC_COLS, start=3):
            val = row[col]
            fmt = "0.0%" if ci in pct_indices else "0.00"
            data_cell(ws, ri, ci, val, bg=bg, num_fmt=fmt)

    ws.row_dimensions[1].height = 32
    return ws


# ── Sheet 2: CONTROL (dropdowns) ─────────────────────────────────────────────
def build_control_sheet(wb):
    ws = wb.create_sheet("CONTROL")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("B2:H2")
    t = ws["B2"]
    t.value = "SECTOR MEDIANS  |  Interactive Dashboard Controls"
    t.font  = Font(bold=True, size=16, color=WHITE)
    t.fill  = PatternFill("solid", fgColor=DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 36

    # ── ACROSS YEARS controls (col B-D) ─────────────────────────────────────
    ws.merge_cells("B4:D4")
    h = ws["B4"]
    h.value = "CHART 1: Metric Over Time (by Sector)"
    h.font  = Font(bold=True, size=12, color=WHITE)
    h.fill  = PatternFill("solid", fgColor=MID_BLUE)
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 24

    for label, cell_addr, default in [
        ("Select Sector:",  "C6",  "Basic Materials"),
        ("Select Metric:",  "C8",  "PE Ratio"),
    ]:
        ws[cell_addr.replace("C","B")] = label
        ws[cell_addr.replace("C","B")].font = Font(bold=True, size=11, color=DARK_BLUE)
        c = ws[cell_addr]
        c.value     = default
        c.font      = Font(size=11, bold=True, color=DARK_BLUE)
        c.fill      = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()

    # Dropdown: sector
    dv_sector1 = DataValidation(
        type="list",
        formula1='"' + ",".join(SECTORS) + '"',
        allow_blank=False,
        showDropDown=False,
    )
    dv_sector1.sqref = "C6"
    ws.add_data_validation(dv_sector1)

    # Dropdown: metric
    dv_metric1 = DataValidation(
        type="list",
        formula1='"' + ",".join(METRIC_LABELS) + '"',
        allow_blank=False,
        showDropDown=False,
    )
    dv_metric1.sqref = "C8"
    ws.add_data_validation(dv_metric1)

    # ── ACROSS SECTORS controls (col F-H) ────────────────────────────────────
    ws.merge_cells("F4:H4")
    h2 = ws["F4"]
    h2.value = "CHART 2: All Sectors Snapshot (by Year)"
    h2.font  = Font(bold=True, size=12, color=WHITE)
    h2.fill  = PatternFill("solid", fgColor=MID_BLUE)
    h2.alignment = Alignment(horizontal="center", vertical="center")

    for label, cell_addr, default in [
        ("Select Year:",   "G6",  "2024"),
        ("Select Metric:", "G8",  "PE Ratio"),
    ]:
        ws[cell_addr.replace("G","F")] = label
        ws[cell_addr.replace("G","F")].font = Font(bold=True, size=11, color=DARK_BLUE)
        c = ws[cell_addr]
        c.value     = default
        c.font      = Font(size=11, bold=True, color=DARK_BLUE)
        c.fill      = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()

    # Dropdown: year
    dv_year = DataValidation(
        type="list",
        formula1='"' + ",".join(YEARS) + '"',
        allow_blank=False,
        showDropDown=False,
    )
    dv_year.sqref = "G6"
    ws.add_data_validation(dv_year)

    # Dropdown: metric (chart 2)
    dv_metric2 = DataValidation(
        type="list",
        formula1='"' + ",".join(METRIC_LABELS) + '"',
        allow_blank=False,
        showDropDown=False,
    )
    dv_metric2.sqref = "G8"
    ws.add_data_validation(dv_metric2)

    # Column widths
    for col, w in [("A",3),("B",20),("C",22),("D",3),("E",3),("F",20),("G",22),("H",3)]:
        ws.column_dimensions[col].width = w

    # Instructions
    ws["B11"] = "HOW TO USE:"
    ws["B11"].font = Font(bold=True, size=11, color=DARK_BLUE)
    instructions = [
        "1.  Use the dropdowns above to select sector, year, and metric.",
        "2.  Navigate to the ACROSS_YEARS sheet to see the time-series line chart.",
        "3.  Navigate to the ACROSS_SECTORS sheet to see the sector comparison bar chart.",
        "4.  Charts update automatically when you change a selection above.",
        "",
        "Note: Charts are powered by helper tables on each chart sheet that use",
        "      IFERROR/SUMPRODUCT formulas — no macros or VBA required.",
    ]
    for i, line in enumerate(instructions, start=12):
        ws.cell(row=i, column=2, value=line).font = Font(size=10, color="404040")

    return ws


# ── Sheet 3: ACROSS_YEARS ─────────────────────────────────────────────────────
def build_across_years_sheet(wb, rows):
    """
    Helper table: col A = year (2011-2026), col B = single value for the sector
    chosen in CONTROL!C6 and metric in CONTROL!C8.  The chart binds to this one
    series so it updates live when the user changes a dropdown.
    """
    ws = wb.create_sheet("ACROSS_YEARS")
    ws.sheet_view.showGridLines = False

    # Title area
    ws.merge_cells("A1:D2")
    t = ws["A1"]
    t.value = "Sector Medians Over Time"
    t.font  = Font(bold=True, size=14, color=WHITE)
    t.fill  = PatternFill("solid", fgColor=DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 8

    # Instruction row
    ws["A3"] = "→ Change selections on CONTROL sheet to update this chart"
    ws["A3"].font = Font(italic=True, size=10, color=MID_BLUE)
    ws.merge_cells("A3:D3")

    # Live labels pulled from CONTROL
    ws["A4"] = "Sector:"
    ws["A4"].font = Font(bold=True, size=10)
    ws["B4"] = "=CONTROL!C6"
    ws["B4"].font = Font(bold=True, size=10, color=MID_BLUE)
    ws["C4"] = "Metric:"
    ws["C4"].font = Font(bold=True, size=10)
    ws["D4"] = "=CONTROL!C8"
    ws["D4"].font = Font(bold=True, size=10, color=MID_BLUE)

    TABLE_ROW_START = 7
    n_data  = len(rows)
    # DATA rows span 2 .. n_data+1 (row 1 is header)
    data_last = n_data + 1

    # Header row of helper table
    c_yr  = ws.cell(row=TABLE_ROW_START, column=1, value="Year")
    c_yr.font  = Font(bold=True, size=10, color=DARK_BLUE)
    c_yr.fill  = PatternFill("solid", fgColor=LIGHT_BLUE)
    c_yr.alignment = Alignment(horizontal="center")
    c_yr.border = thin_border()

    header_cell(ws, TABLE_ROW_START, 2, "Value", bg=LIGHT_BLUE, fg=DARK_BLUE, size=10)

    for ri, year in enumerate(YEARS, start=TABLE_ROW_START + 1):
        bg = WHITE if ri % 2 == 0 else GREY_BG

        c = ws.cell(row=ri, column=1, value=int(year))
        c.font      = Font(bold=True, size=10)
        c.fill      = PatternFill("solid", fgColor=LIGHT_BLUE)
        c.alignment = Alignment(horizontal="center")
        c.border    = thin_border()

        # Sector from CONTROL!C6 (text comparison with DATA col A).
        # Year is embedded as a literal integer — matches the integer stored in DATA col B.
        # Metric column resolved dynamically via MATCH on DATA row 1.
        formula = (
            f'=IFERROR(SUMPRODUCT('
            f'(DATA!$A$2:$A${data_last}=CONTROL!$C$6)*'
            f'(DATA!$B$2:$B${data_last}={year})*'
            f'INDEX(DATA!$C$2:$Q${data_last},0,'
            f'MATCH(CONTROL!$C$8,DATA!$C$1:$Q$1,0))'
            f'),"")'
        )
        c2 = ws.cell(row=ri, column=2, value=formula)
        c2.fill      = PatternFill("solid", fgColor=bg)
        c2.alignment = Alignment(horizontal="center")
        c2.border    = thin_border()
        c2.number_format = "0.00"

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 16

    # ── Line chart — single series, one line per selected sector ────────────
    chart = LineChart()
    chart.title    = None
    chart.style    = 10
    chart.height   = 14
    chart.width    = 30
    chart.grouping = "standard"
    chart.smooth   = True

    cats = Reference(ws,
                     min_col=1, max_col=1,
                     min_row=TABLE_ROW_START + 1,
                     max_row=TABLE_ROW_START + len(YEARS))
    chart.set_categories(cats)

    # min_row=TABLE_ROW_START includes the "Value" header → series label in legend
    values = Reference(ws,
                       min_col=2, max_col=2,
                       min_row=TABLE_ROW_START,
                       max_row=TABLE_ROW_START + len(YEARS))
    series = Series(values, title_from_data=True)
    series.smooth = True
    series.graphicalProperties.line.width     = 25000   # ~2.5 pt
    series.graphicalProperties.line.solidFill = MID_BLUE
    chart.series.append(series)

    chart.legend.position  = "b"
    chart.plot_area.layout = None

    chart_row = TABLE_ROW_START + len(YEARS) + 2
    ws.add_chart(chart, f"A{chart_row}")

    return ws


# ── Sheet 4: ACROSS_SECTORS ──────────────────────────────────────────────────
def build_across_sectors_sheet(wb, rows):
    """
    Helper table: col A = sector name, col B = value for the year chosen in
    CONTROL!G6 and metric in CONTROL!G8.
    Bar chart shows all sectors side-by-side for the selected year.

    Key fix vs original: year comparison uses VALUE(CONTROL!$G$6) to coerce the
    text from the dropdown into a number matching the integer stored in DATA col B.
    """
    ws = wb.create_sheet("ACROSS_SECTORS")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:F2")
    t = ws["A1"]
    t.value = "Sector Medians — Cross-Sectional Snapshot"
    t.font  = Font(bold=True, size=14, color=WHITE)
    t.fill  = PatternFill("solid", fgColor=DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 8

    ws["A3"] = "→ Change selections on CONTROL sheet to update this chart"
    ws["A3"].font = Font(italic=True, size=10, color=MID_BLUE)
    ws.merge_cells("A3:F3")

    ws["A4"] = "Year:"
    ws["A4"].font = Font(bold=True, size=10)
    ws["B4"] = "=CONTROL!G6"
    ws["B4"].font = Font(bold=True, size=10, color=MID_BLUE)
    ws["C4"] = "Metric:"
    ws["C4"].font = Font(bold=True, size=10)
    ws["D4"] = "=CONTROL!G8"
    ws["D4"].font = Font(bold=True, size=10, color=MID_BLUE)

    TABLE_ROW_START = 7
    n_data = len(rows)
    data_last = n_data + 1

    # Headers
    c_hdr1 = ws.cell(row=TABLE_ROW_START, column=1, value="Sector")
    c_hdr1.font  = Font(bold=True, size=10, color=DARK_BLUE)
    c_hdr1.fill  = PatternFill("solid", fgColor=LIGHT_BLUE)
    c_hdr1.border = thin_border()
    c_hdr2 = ws.cell(row=TABLE_ROW_START, column=2, value="Value")
    c_hdr2.font  = Font(bold=True, size=10, color=DARK_BLUE)
    c_hdr2.fill  = PatternFill("solid", fgColor=LIGHT_BLUE)
    c_hdr2.border = thin_border()

    for ri, sector in enumerate(SECTORS, start=TABLE_ROW_START + 1):
        bg = WHITE if ri % 2 == 0 else GREY_BG
        c1 = ws.cell(row=ri, column=1, value=sector)
        c1.fill      = PatternFill("solid", fgColor=bg)
        c1.alignment = Alignment(horizontal="left")
        c1.border    = thin_border()

        # VALUE() coerces the text year from the dropdown to an integer so it
        # matches the integer stored in DATA col B — without this the SUMPRODUCT
        # returns 0 for every row and the chart appears empty.
        formula = (
            f'=IFERROR(SUMPRODUCT('
            f'(DATA!$A$2:$A${data_last}="{sector}")*'
            f'(DATA!$B$2:$B${data_last}=VALUE(CONTROL!$G$6))*'
            f'INDEX(DATA!$C$2:$Q${data_last},0,'
            f'MATCH(CONTROL!$G$8,DATA!$C$1:$Q$1,0))'
            f'),"")'
        )
        c2 = ws.cell(row=ri, column=2, value=formula)
        c2.fill      = PatternFill("solid", fgColor=bg)
        c2.alignment = Alignment(horizontal="center")
        c2.border    = thin_border()
        c2.number_format = "0.00"

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14

    # ── Bar chart ─────────────────────────────────────────────────────────────
    chart = BarChart()
    chart.type     = "col"
    chart.style    = 10
    chart.title    = None
    chart.height   = 14
    chart.width    = 22
    chart.grouping = "clustered"
    chart.overlap  = 0

    cats = Reference(ws,
                     min_col=1, max_col=1,
                     min_row=TABLE_ROW_START + 1,
                     max_row=TABLE_ROW_START + len(SECTORS))
    values = Reference(ws,
                       min_col=2, max_col=2,
                       min_row=TABLE_ROW_START,        # include header row for series label
                       max_row=TABLE_ROW_START + len(SECTORS))

    series = Series(values, title_from_data=True)
    chart.series.append(series)
    chart.set_categories(cats)
    chart.legend.position = "b"

    chart_row = TABLE_ROW_START + len(SECTORS) + 2
    ws.add_chart(chart, f"A{chart_row}")

    return ws


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("Loading data from DB...")
    rows = load_data()
    print(f"  {len(rows)} rows loaded ({len(SECTORS)} sectors × {len(YEARS)} years)")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Building DATA sheet...")
    build_data_sheet(wb, rows)

    print("Building CONTROL sheet...")
    ws_ctrl = build_control_sheet(wb)

    print("Building ACROSS_YEARS sheet...")
    build_across_years_sheet(wb, rows)

    print("Building ACROSS_SECTORS sheet...")
    build_across_sectors_sheet(wb, rows)

    # Set CONTROL as the active sheet (shown first on open)
    wb.active = wb["CONTROL"]

    print(f"Saving to {OUT_PATH} ...")
    wb.save(OUT_PATH)
    print("Done.")


if __name__ == "__main__":
    main()
